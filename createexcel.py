import plotly
import plotly.graph_objects as go
from pandas import read_excel
from openpyxl import load_workbook
from sys import argv


annotation_template = go.layout.Template()

config = dict({'scrollZoom': True,'displaylogo': False,'modeBarButtonsToAdd':['drawline',
                                        'drawopenpath',
                                        'drawclosedpath',
                                        'drawcircle',
                                        'drawrect',
                                        'eraseshape'
                                       ]})


droppedFile = argv

rawfile = 'concat.xlsx'
title = 'input'

with open(rawfile) as sheeter:
    wb = load_workbook(filename = rawfile)
    sheets = wb.sheetnames 

fig = go.Figure()

for sheet in sheets:
    try:
        df = read_excel(rawfile,sheet_name=sheet,engine='openpyxl',index_col=None, header=None)
        df = df.interpolate()
        df = df.fillna(0.0)
        # Create traces
        
        fig.add_trace(go.Scatter(x=df[0], y=df[1],
                            mode='lines',
                            name=sheet,connectgaps=True))
    except:
        pass

fig.update_layout(template=annotation_template,title=title,
                xaxis_title='Time',
                yaxis_title='Temperature (degrees C)'
                )

fig.add_layout_image(
dict(
    source= r"https://iconape.com/wp-content/files/mo/195650/svg/195650.svg",
    xref="paper", yref="paper",
    x=1, y=1.0,
    sizex=0.1, sizey=0.1,
    xanchor="right", yanchor="bottom"
)
)

fig.show(config=config)