from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from sys import argv

wb_concat = Workbook()

concat_filename = 'Dados_All.xlsx'
droppedFile = argv

for file in droppedFile[1:]:
    wb = load_workbook(file)
    for sheet in wb.sheetnames:
        ws1 = wb_concat.create_sheet(title=sheet) 
        ws = wb[sheet]
        for x in ws.values:
            ws1.append(x)

Sheet1 = wb_concat['Sheet']
wb_concat.remove(Sheet1)
wb_concat.save(filename = concat_filename)